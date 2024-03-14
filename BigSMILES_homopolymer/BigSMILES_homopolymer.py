import pandas as pd
import numpy as np
import time

import csv
from openpyxl import load_workbook

import re
from collections import Counter

from rdkit import Chem
from rdkit import RDLogger


###         Sunho Choi (schoi_@korea.ac.kr)
###         Korea University, School of Electrical Engineering
###         24.03.13
###         version : 1.0.3
###
###         Article:
###         An automated BigSMILES conversion workflow and dataset for homopolymeric macromolecules
###
###
###
### Additional notes in 24.03.13 :
###
### Note that this SMILES to BigSMILES converter operates to convert to the BigSMILES version 1.0 notation, as defined in original BigSMILES paper(2019).
### To convert BigSMILES to version 1.1, use the class version_conveter defined from line 725 onwards to convert BigSMILES created with 1.0.
###


class SMILES2BigSMILES:

    ### initialization
    def __init__(self):
        self.SMILES_data = None
        self.num_data = 0

        self.BigSMILES_data = None

        self.nums = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '0']

        RDLogger.DisableLog('rdApp.*')

        self.check_unit_detail = [
            r'N\d[()]*(c\d?\d?[()]*)*\(=O\)[()]*(c\d?\d?[()]*)*=O[()]*([cofs]\d?\d?[()]*)*\(=O\)[()]*(c\d?\d?[()]*)*N[()]*(c\d?\d?[()]*)*=O[()]*(c\d?\d?[()]*)*\)',
            r'C\(=O\)(c\d?\d?[()]*)+OC\(=O\)(c\d?\d?[()]*)+C\(=O\)',
            r'C\(=O\)(c\d?\d?[()]*)+C\(=O\)',
            r'c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?C\(=O\)[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\)?',
            r'N[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?C\d[()]?[()]?OC\(=O\)c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\d[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?N\)?',
            r'O[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?C\d[()]?[()]?OC\(=O\)c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\d[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?O\)?',
            r'c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?O[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?O[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\)?',
            r'Oc\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?S\(=O\)\(=O\)[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\)?O',
            r'O[()]*(c\d?\d?[()]*)+(F[()]*)*(Cl[()]*)*(F[()]*)*(Cl[()]*)*(c\d?\d?[()]*)*(F[()]*)*(Cl[()]*)*(F[()]*)*(Cl[()]*)*(c\d?\d?[()]*)+O',
            r'O[()]*(c\d?\d?[()]*)+Cl[()]*(c\d?\d?[()]*)*Cl[()]*(c\d?\d?[()]*)*Cl[()]*(c\d?\d?[()]*)*Cl[()]*(c\d?\d?[()]*)+O',
            r'O[()]*(c\d?\d?[()]*)+Cl[()]*(c\d?\d?[()]*)*Cl[()]*(c\d?\d?[()]*)+O',
            r'O[()]*(c\d?\d?[()]*)+Br[()]*(c\d?\d?[()]*)*Br[()]*(c\d?\d?[()]*)*Br[()]*(c\d?\d?[()]*)*Br[()]*(c\d?\d?[()]*)+O',
            r'O[()]*(c\d?\d?[()]*)+Br[()]*(c\d?\d?[()]*)*Br[()]*(c\d?\d?[()]*)+O',
            r'O[()]*(c\d?\d?[()]*)+C\(=\(Cl\)Cl[()]*(c\d?\d?[()]*)+O',
            r'O[()]*(c\d?\d?[()]*)+(Cl)?S?[()]*(c\d?\d?[()]*)+O',
            r'c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?S\(=O\)\(=O\)[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\)?',
            r'c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d[()]?O[()]?c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\)?',
            r'c\d[()]?c[()]?c[()]?c[()]?c[()]?c\d\)?',
            r'O(c\d?\d?[()]*)+O(c\d?\d?[()]*)+O',  ## 22.06.16.
            r'O(c\d?\d?[()]*)+O',
            r'c\d[()]?c[()]?c[()]?c[()]?n[()]?n\d\)',
            r'C\(=O\)(c\d?\d?[()]*)+O(c\d?\d?[()]*)+O(c\d?\d?[()]*)+O(c\d?\d?[()]*)+C\(=O\)',
            r'C\(=O\)(c\d?\d?[()]*)+O(c\d?\d?[()]*)+C\(=O\)',
        ]

        self.check_unit_full_match = [
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+S\(=O\)\(=O\)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+O(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',
            ## 22.06.16.
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+S\(=O\)\(=O\)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+O(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',  ## 22.06.16.
            r'(c\d?\d?[()]*)+S\(=O\)\(=O\)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+C\(=O\)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+C\(=O\)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+C\(=O\)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+O(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+O(c\d?\d?[()]*)+(n\d?\d?[()]*)(c\d?\d?[()]*)+',
            r'(c\d?\d?[()]*)+O(c\d?\d?[()]*)+',
            r'N(c\d?\d?[()]*)+O(c\d?\d?[()]*)+N',
            r'N(c\d?\d?[()]*)+N',
            r'(c\d?c\d?\[()]*)+',
            r'C\(=O\)',
            r'CC\(C\(C\)=C\)',  ### Common Repeat Units from BigSMILES Github
            r'C\\C=C\(C\)\/C',
            r'C\/C=C\(C\)\/C',
            r'CC\(OC\(=O\)C\)',
            r'OC\(C\)C\(=O\)',
            r'C\(=O\)CCCCC\(=O\)NCCCCCCN',
            r'CC=CC',
            r'C\\C=C\/C',
            r'C\/C=C\/C',
            r'CC\(Cl\)=CC',
            r'CC\(Cl\)',
            r'C\(F\)\(F\)C\(F\)\(F\)',
            r'CC\(F\)\(F\)',
            r'CC\(C\)\(C\(=O\)OC\)',
            r'CC\(C\(=O\)OC\)',
            r'CC\(C\)\(C\(=O\)O\)',
            r'CC\(C\(=O\)O\)',
            r'CC\(C\)\(C\(=O\)\[O-\]\)',
            r'CC\(C\(=O\)\[O-\]\)',
            r'CC\(C\(=O\)OCCCC\)',
            r'CC\(C\(=O\)OCC\)',
            r'CC\(C\(=O\)OCC\(CC\)CCCC\)',
            r'CC\(C\(=O\)N\)',
            r'CC\(C#N\)',
            r'CC\(C\(=O\)NC\(C\)C\)',
            r'CC\(C\(=O\)N\(C\)C\)',
            r'CC\(c\dccc\(Cl\)cc\d\)',
            r'CC\(c\dc\(Cl\)cccc\d\)',
            r'CC\(c\dccc\(CCl\)cc\d\)',
            r'CC\(c\dccc\(S\(=O\)\(=O\)O\)cc\d\)',
            r'CC\(c\dccc\(S\(=O\)\(=O\)\[O-\]\)cc\d\)',
            r'CC\(c\dccncc\d\)',
            r'CC\(c\dncccc\d\)',
            r'CC\(C\)O',
            r'OCC\(=O\)',
            r'OCCCCCC\(=O\)',
            r'NCCCCCC\(=O\)',
            r'Cc\dccc\(cc\d\)C',
            r'Cc\dccccc\(c\d\)C',
            r'C\(S\d\)=CC=C\d',
            r'c\(cc\d\)ccc\dC=C',
            r'C\(S\d\)=C\dOCCOC\d=C\d',
            r'C\(N\d\)=CC=C\d',
            r'c\(cc\d\)ccc\dS',
            r'C=C',
            r'c\(cc\d\)ccc\dN',
            r'c\(cc\d\)ccc\dN=C\(C=C\d\)C=CC\d=N',
        ]

        self.check_unit_full_match.extend(self.check_unit_detail)

    ### data loading functions
    ###
    ### Data_Load_xlsx() : loading xlsx type file. requires SMILES header in the input file.
    ### Data_Load_csv() : loading csv type file. requires SMILES header in the input file.
    ### Data_Load_txt() : loading txt type file. only requires SMILES with no header. the user can set how much to read the line when the input file is too large.
    ###
    ### Input SMILES column must be a form of repeating unit of homopolymers.
    ###

    def Data_Load_xlsx(self, filename, data_only=True, SMILES='SMILES'):
        load_wb = load_workbook(filename, data_only)
        load_ws = load_wb.active
        data = pd.DataFrame([[i.value for i in j] for j in load_ws.rows])
        header = data.iloc[0]
        data = data[1:]
        data.rename(columns=header, inplace=True)
        data.dropna(subset=[SMILES], inplace=True)
        data.reset_index(drop=True, inplace=True)

        self.SMILES_data = data[SMILES]

        self.num_data = len(self.SMILES_data)

    def Data_Load_csv(self, filename, SMILES='SMILES', encoding='utf-8'):
        data = pd.read_csv(filename, header=0, encoding=encoding)
        data.dropna(subset=[SMILES], inplace=True)
        data.reset_index(drop=True, inplace=True)

        self.SMILES_data = data[SMILES]

        self.num_data = len(self.SMILES_data)

    def Data_Load_txt(self, filename, N=100000, smiles_len=85):
        ### No header, only SMILES
        ### SMILES len under 85

        head = []
        with open(filename) as myfile:
            for x in range(N):
                line = myfile.readline()
                if len(line) < smiles_len:
                    head.append(line)
        print('total smiles under length limit: ' + str(len(head)))
        data = pd.DataFrame({"SMILES": head})
        head = None

        data.dropna(inplace=True)
        data.reset_index(drop=True, inplace=True)

        self.SMILES_data = data['SMILES']

        self.num_data = len(self.SMILES_data)

    ### Since it is a necessary part of function implementation, it is not called in actual testing.

    def assign_bracket_hierarchy(self, smiles):
        ### Code needed to convert SMILES' asterisk to both ends
        bracket_pairs = {'(': ')'}
        hierarchy_list = [0] * len(smiles)
        stack = []
        current_hierarchy = 1

        for i, char in enumerate(smiles):
            if char in bracket_pairs.keys():
                stack.append((char, current_hierarchy))
                hierarchy_list[i] = current_hierarchy
                current_hierarchy += 1
            elif char in bracket_pairs.values():
                if stack:
                    start_bracket, start_hierarchy = stack.pop()
                    hierarchy_list[i] = -start_hierarchy
            elif char in '*':
                hierarchy_list[i] = 'A'

        return hierarchy_list

    def switching(self, smiles):
        ### Code needed to convert SMILES' asterisk to both ends
        alist = self.assign_bracket_hierarchy(smiles)
        star = smiles.rfind('*')
        tmp = smiles

        if star == len(smiles): return smiles
        while True:
            if star == len(smiles) - 1:
                return tmp
            check = 0
            not_large_x = None
            not_large_y = None

            for i in range(star - 1, 0, -1):
                check = check + alist[i]
                if check > 0: break

            if check > 0:
                small_x = i
                small_y = alist.index(-check)
                check2 = 0
                for i in range(small_y + 1, len(smiles)):
                    check2 = check2 + alist[i]
                    if check2 < 0: break

                if alist[small_y + 1] > 0:
                    not_large_x = alist.index(alist[small_y + 1])
                    not_large_y = alist.index(-alist[small_y + 1])

                if check2 >= 0:
                    large_x = 0
                    large_y = len(smiles)
                else:
                    large_x = alist.index(-check2)
                    large_y = i

                tmp2 = re.sub('(\[[a-zA-Z]?)\d([a-zA-Z]?\])', '\\1A\\2', tmp)
                tmp2 = re.sub('(\[[a-zA-Z]?)\d\d([a-zA-Z]?\])', '\\1AA\\2', tmp2)
                tmp2 = re.sub('(\[[a-zA-Z]?)\d\d\d([a-zA-Z]?\])', '\\1AAA\\2', tmp2)

                counts = Counter(re.sub(r'\[[^()]*\]|\s-\s.*', '', tmp2[small_x:small_y]))
                num_switches = []
                for nu in self.nums:
                    if counts[nu] % 2 != 0: num_switches.append(nu)

                if len(num_switches):
                    counts2 = Counter(tmp2[small_y:large_y])
                    for num_switch in num_switches:
                        if counts2[num_switch]:
                            ind_x = len(tmp2[:small_x]) + tmp2[small_x:small_y].find(str(num_switch))  ## ring closure
                            ind_y = tmp2[:small_x].rfind(str(num_switch))
                            res = 0
                            counts3 = Counter(tmp2[:large_y])
                            for nu in self.nums:
                                if counts3[nu] == 0:
                                    res = nu
                                    break
                            if res == 0: res = 15
                            tmp = list(tmp)
                            tmp[ind_x] = res
                            tmp[ind_y] = res
                            tmp = ''.join(tmp)

                if not_large_x:
                    tmp = tmp[:large_x] + tmp[large_x:small_x + 1] + tmp[not_large_y + 1:large_y] + tmp[
                                                                                                    small_y:not_large_y + 1] + tmp[
                                                                                                                               small_x + 1:small_y] + tmp[
                                                                                                                                                      large_y:]
                elif check2 >= 0:
                    tmp = tmp[:large_x] + tmp[large_x:small_x + 1] + tmp[small_y + 1:large_y] + ')' + tmp[
                                                                                                      small_x + 1:small_y]
                    return tmp
                else:
                    tmp = tmp[:large_x] + tmp[large_x:small_x + 1] + tmp[small_y + 1:large_y + 1] + tmp[
                                                                                                    small_x + 1:small_y + 1] + tmp[
                                                                                                                               large_y + 1:]

                alist = self.assign_bracket_hierarchy(tmp)
                star = tmp.rfind('*')

            else:
                ### fail if asterisks are not located at both ends
                return None

    ### function to convert SMILES to BigSMILES

    def Converting(self, resultfile='result', move_parallel=-1):
        ### SMILES to BigSMILES converting function
        ### for file input
        ### Previously, data loading functions should be used.
        ###
        ### Parameter Description:
        ### resultfile :    File path and name to save the converted data
        ###                 The processing number and CSV name are automatically added at the end.
        ### move_parallel : Used during the search for two reactive end groups within the same SMILES representation.
        ###                 default = -1 (searching all ordering)
        ###                 For integers greater than or equal to 0, reorder by that number of times

        self.BigSMILES_data = []
        kk = 0

        for i in range(self.num_data):

            print('process : ', i, ' / ', self.num_data)

            if i - kk == 100000:
                aa = pd.DataFrame(self.BigSMILES_data)
                aa.to_csv(resultfile + str(kk) + '.csv')
                kk = i
                self.BigSMILES_data = []
                aa = None

            tmp = Chem.MolToSmiles(Chem.MolFromSmiles(self.SMILES_data[i]))
            tmp = re.sub(r'(\[[^\[^\]^\-]*)-([^\[^\]^\-]*\])', '\\1:\\2', tmp)
            tmp = re.sub('-', '', tmp)
            SMILES = re.sub(':', '-', tmp)

            br = self.switching(SMILES)

            if br:
                SMILES = br
            else:
                self.BigSMILES_data.append('')
                continue

            temporary_SMILES = [SMILES.strip('*')]

            SMILES_delete1 = re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES.strip('*'))
            SMILES_delete = re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES_delete1)
            temporary_SMILES.append(SMILES_delete1)
            temporary_SMILES.append(SMILES_delete)

            while True:
                if len(SMILES_delete) == len(re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES_delete)): break
                SMILES_delete = re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES_delete)
                temporary_SMILES.append(SMILES_delete)

            chain_check = Counter(SMILES_delete)

            for nu in self.nums:
                if chain_check[nu] % 2 == 1: break

            ### Main chain checking
            if re.sub(r'[Cc0-9=\\/#]', '', SMILES_delete) == '':  ## carbon
                if int(nu) > 0:
                    first_num = SMILES_delete.index(nu)
                    bre = False
                    for j in range(1, len(temporary_SMILES)):
                        if temporary_SMILES[-j].rfind(nu) != first_num:
                            if re.sub(r'[Cc0-9=\\/#()]', '',
                                      temporary_SMILES[-j][:temporary_SMILES[-j].rfind(nu)]) == '':
                                dollar = re.sub(r'\*', '$', SMILES)
                                self.BigSMILES_data.append('{' + dollar + '}')
                                bre = True
                                break
                    if bre: continue
                else:
                    dollar = re.sub(r'\*', '$', SMILES)
                    self.BigSMILES_data.append('{' + dollar + '}')
                    continue


            elif re.sub(r'[Si\[\]]', '', SMILES_delete) == '':  ## silane
                dollar = re.sub(r'\*', '$', SMILES)
                self.BigSMILES_data.append('{' + dollar + '}')
                continue

            SMILES = SMILES.strip('*')

            if ('CCO' == SMILES_delete1 and SMILES != SMILES_delete1):
                ### checking if the input SMILES is "Propylene" and AB type bond <>
                ### more details in BigSMILES paper suppliementary information.
                first = '<' + SMILES + '>'
                second = '<' + SMILES[1:-1] + SMILES[0] + SMILES[-1] + '>'

                self.BigSMILES_data.append('{' + first + ',' + second + '}')


            else:
                ### checking if we can find two reactive groups.
                ### if we cannot found anything, Simplified BigSMILES representation will be out.

                moving_atoms = [SMILES]
                if move_parallel > 0:
                    moving_atoms.extend([SMILES[p:] + SMILES[0:p] for p in range(move_parallel)])
                    moving_atoms.extend([SMILES[-p:] + SMILES[0:-p] for p in range(move_parallel)])
                elif move_parallel == 0:
                    pass
                else:
                    moving_atoms.extend([SMILES[p:] + SMILES[0:p] for p in range(int(len(SMILES) / 2))])
                    moving_atoms.extend([SMILES[-p:] + SMILES[0:-p] for p in range(int(len(SMILES) / 2))])

                status = False

                for check in range(len(self.check_unit_detail)):

                    if status:
                        if tmp3.regs[0][0] == 0:
                            break
                        elif tmp3.regs[-1][1] == len(moved_smiles):
                            break
                        else:
                            status = False
                    if status: break

                    for moved_smiles in moving_atoms:
                        status = False

                        if re.search(r'\)', moved_smiles):
                            if moved_smiles.rfind(')') < moved_smiles.rfind('('):
                                continue
                            elif moved_smiles.find(')') < moved_smiles.find('('):
                                continue
                            elif moved_smiles.find(']') < moved_smiles.find('['):
                                continue

                        if Chem.MolFromSmiles('*' + moved_smiles + '*') == None: continue

                        tmp3 = re.match(self.check_unit_detail[check].lower(), moved_smiles.lower())

                        if tmp3:
                            if tmp3.regs[0][0] == 0:
                                tmp4 = moved_smiles[tmp3.regs[0][1]:].lower()
                                status = True

                            elif tmp3.regs[-1][1] == len(moved_smiles):
                                tmp4 = moved_smiles[:tmp3.regs[-1][0]].lower()
                                status = True

                            point1 = False
                            if status:
                                for check2 in range(len(self.check_unit_full_match)):
                                    if tmp4 == None: break
                                    if re.match(self.check_unit_full_match[check2].lower(), tmp4):
                                        if re.match(self.check_unit_full_match[check2].lower(), tmp4).regs[0][1] == len(
                                                tmp4):
                                            point1 = True
                                            counts = Counter(tmp4)
                                            if counts['('] != counts[')']: point1 = False
                                            for nu in self.nums:
                                                if counts[nu] % 2 != 0: point1 = False
                                    if point1: break

                                if point1:
                                    status = True
                                    break
                                else:
                                    status = False

                if status:
                    if tmp3.regs[0][0] == 0:
                        self.BigSMILES_data.append(
                            '{<' + moved_smiles[:tmp3.regs[0][1]] + '<,>' + moved_smiles[
                                                                            tmp3.regs[0][1]:] + '>}')
                    elif tmp3.regs[-1][1] == len(moved_smiles):
                        self.BigSMILES_data.append(
                            '{<' + moved_smiles[:tmp3.regs[-1][0]] + '<,>' + moved_smiles[
                                                                             tmp3.regs[-1][0]:] + '>}')
                else:
                    self.BigSMILES_data.append('{<' + SMILES + '>}')

        if kk != i:
            aa = pd.DataFrame(self.BigSMILES_data)
            aa.to_csv(resultfile + str(kk) + '.csv')
            kk = i
            self.BigSMILES_data = []
            aa = None

    def Converting_single(self, SMILES, move_parallel=-1):
        ### SMILES to BigSMILES converting function
        ### for single SMILES input
        ###
        ### Parameter Description:
        ### SMILES :        One repeating unit SMILES
        ###                 Must be a repeating unit of the Homopolymer
        ### move_parallel : Used during the search for two reactive end groups within the same SMILES representation.
        ###                 default = -1 (searching all ordering)
        ###                 For integers greater than or equal to 0, reorder by that number of times

        tmp = Chem.MolToSmiles(Chem.MolFromSmiles(SMILES))
        tmp = re.sub(r'(\[[^\[^\]^\-]*)-([^\[^\]^\-]*\])', '\\1:\\2', tmp)
        tmp = re.sub('-', '', tmp)
        SMILES = re.sub(':', '-', tmp)

        br = self.switching(SMILES)

        if br:
            SMILES = br
        else:
            return 0

        temporary_SMILES = [SMILES.strip('*')]

        SMILES_delete1 = re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES.strip('*'))
        SMILES_delete = re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES_delete1)
        temporary_SMILES.append(SMILES_delete1)
        temporary_SMILES.append(SMILES_delete)

        while True:
            if len(SMILES_delete) == len(re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES_delete)): break
            SMILES_delete = re.sub(r'\([^()]*\)|\s-\s.*', '', SMILES_delete)
            temporary_SMILES.append(SMILES_delete)

        chain_check = Counter(SMILES_delete)

        for nu in self.nums:
            if chain_check[nu] % 2 == 1: break

        ## Main chain checking
        if re.sub(r'[Cc0-9=\\/#]', '', SMILES_delete) == '':  ## carbon
            if int(nu) > 0:
                first_num = SMILES_delete.index(nu)
                bre = False
                for j in range(1, len(temporary_SMILES)):
                    if temporary_SMILES[-j].rfind(nu) != first_num:
                        if re.sub(r'[Cc0-9=\\/#()]', '', temporary_SMILES[-j][:temporary_SMILES[-j].rfind(nu)]) == '':
                            dollar = re.sub(r'\*', '$', SMILES)
                            return '{' + dollar + '}'
            else:
                dollar = re.sub(r'\*', '$', SMILES)
                return '{' + dollar + '}'


        elif re.sub(r'[Si\[\]]', '', SMILES_delete) == '':  ## silane
            dollar = re.sub(r'\*', '$', SMILES)
            return '{' + dollar + '}'

        SMILES = SMILES.strip('*')

        if ('CCO' == SMILES_delete1 and SMILES != SMILES_delete1):
            ### checking if the input SMILES is "Propylene" and AB type bond <>
            ### more details in BigSMILES paper suppliementary information.
            first = '<' + SMILES + '>'
            second = '<' + SMILES[1:-1] + SMILES[0] + SMILES[-1] + '>'

            return '{' + first + ',' + second + '}'


        else:
            ### checking if we can find reactive groups.
            ### if we cannot found anything, Simplified BigSMILES representation will be return.

            moving_atoms = [SMILES]
            if move_parallel > 0:
                moving_atoms.extend([SMILES[p:] + SMILES[0:p] for p in range(move_parallel)])
                moving_atoms.extend([SMILES[-p:] + SMILES[0:-p] for p in range(move_parallel)])
            elif move_parallel == 0:
                pass
            else:
                moving_atoms.extend([SMILES[p:] + SMILES[0:p] for p in range(int(len(SMILES) / 2))])
                moving_atoms.extend([SMILES[-p:] + SMILES[0:-p] for p in range(int(len(SMILES) / 2))])

            status = False

            for check in range(len(self.check_unit_detail)):

                if status:
                    if tmp3.regs[0][0] == 0:
                        break
                    elif tmp3.regs[-1][1] == len(moved_smiles):
                        break
                    else:
                        status = False
                if status: break

                for moved_smiles in moving_atoms:
                    status = False

                    if re.search(r'\)', moved_smiles):
                        if moved_smiles.rfind(')') < moved_smiles.rfind('('):
                            continue
                        elif moved_smiles.find(')') < moved_smiles.find('('):
                            continue
                        elif moved_smiles.find(']') < moved_smiles.find('['):
                            continue

                    if Chem.MolFromSmiles('*' + moved_smiles + '*') == None: continue

                    tmp3 = re.match(self.check_unit_detail[check].lower(), moved_smiles.lower())

                    if tmp3:
                        if tmp3.regs[0][0] == 0:
                            tmp4 = moved_smiles[tmp3.regs[0][1]:].lower()
                            status = True

                        elif tmp3.regs[-1][1] == len(moved_smiles):
                            tmp4 = moved_smiles[:tmp3.regs[-1][0]].lower()
                            status = True

                        point1 = False
                        if status:
                            for check2 in range(len(self.check_unit_full_match)):
                                if tmp4 == None: break
                                if re.match(self.check_unit_full_match[check2].lower(), tmp4):
                                    if re.match(self.check_unit_full_match[check2].lower(), tmp4).regs[0][1] == len(
                                            tmp4):
                                        point1 = True
                                        counts = Counter(tmp4)
                                        if counts['('] != counts[')']: point1 = False
                                        for nu in self.nums:
                                            if counts[nu] % 2 != 0: point1 = False
                                if point1: break

                            if point1:
                                status = True
                                break
                            else:
                                status = False

            if status:
                if tmp3.regs[0][0] == 0:
                    return '{<' + moved_smiles[:tmp3.regs[0][1]] + '<,>' + moved_smiles[
                                                                           tmp3.regs[0][1]:] + '>}'
                elif tmp3.regs[-1][1] == len(moved_smiles):
                    return '{<' + moved_smiles[:tmp3.regs[-1][0]] + '<,>' + moved_smiles[
                                                                            tmp3.regs[-1][0]:] + '>}'
            else:
                return '{<' + SMILES + '>}'



class BigSMILES2SMILES:
    def __init__(self):
        self.SMILES_data = None
        self.num = 0

        self.BigSMILES_data = None

    def Data_Load_xlsx(self, filename, data_only=True, BigSMILES='BigSMILES'):
        load_wb = load_workbook(filename, data_only)
        load_ws = load_wb.active
        data = pd.DataFrame([[i.value for i in j] for j in load_ws.rows])
        header = data.iloc[0]
        data = data[1:]
        data.rename(columns=header, inplace=True)
        data.dropna(subset=[BigSMILES], inplace=True)
        data.reset_index(drop=True, inplace=True)

        self.BigSMILES_data = data[BigSMILES]

        self.num = len(self.BigSMILES_data)

    def Data_Load_csv(self, filename, BigSMILES='BigSMILES', encoding='utf-8'):
        ## TXT format files can also be imported
        data = pd.read_csv(filename, header=0, encoding=encoding)
        data.dropna(subset=[BigSMILES], inplace=True)
        data.reset_index(drop=True, inplace=True)

        self.BigSMILES_data = data[BigSMILES]

        self.num = len(self.BigSMILES_data)

    def Converting(self, resultfile='result'):
        self.SMILES_data = []
        kk = 0

        for i in range(self.num):

            print('process : ', i, ' / ', self.num)

            BigSMILES = self.BigSMILES_data[i]

            if '>,<' in BigSMILES:
                ind_x = BigSMILES.rfind(',')
                BigSMILES = BigSMILES[:ind_x]

            for j in ['<', '>', '{', '}', '\$', ',', ' ']:
                BigSMILES = re.sub(j, '', BigSMILES)

            BigSMILES = '*' + BigSMILES + '*'

            BigSMILES = Chem.MolToSmiles(Chem.MolFromSmiles(BigSMILES))

            self.SMILES_data.append(BigSMILES)

            if i - kk == 100000:
                aa = pd.DataFrame(self.SMILES_data)
                aa.to_csv(resultfile + str(kk) + '.csv')
                kk = i
                self.SMILES_data = []
                aa = None

        if kk != i:
            aa = pd.DataFrame(self.SMILES_data)
            aa.to_csv(resultfile + str(kk) + '.csv')
            kk = i
            self.SMILES_data = []
            aa = None

    
    def Converting_single(self, BigSMILES):

        if '>,<' in BigSMILES:
            ind_x = BigSMILES.rfind(',')
            BigSMILES = BigSMILES[:ind_x]

        for j in ['<', '>', '{', '}', '\$', ',', ' ']:
            BigSMILES = re.sub(j, '', BigSMILES)

        BigSMILES = '*' + BigSMILES + '*'

        BigSMILES = Chem.MolToSmiles(Chem.MolFromSmiles(BigSMILES))

        return BigSMILES





class version_converter:
    def zero2one(self, BigSMILES_ver10):
        tmp = BigSMILES_ver10
        tmp = re.sub('<,>', '[<],[>]', tmp)
        tmp = re.sub('>,<', '[>],[<]', tmp)
        tmp = re.sub('{<', '{[][<]', tmp)
        tmp = re.sub('>}', '[>][]}', tmp)
        tmp = re.sub('{$', '{[][$]', tmp)
        tmp = re.sub('$}', '[$][]}', tmp)
        return tmp
    
    
    def one2zero(self, BigSMILES_ver11):
        tmp = BigSMILES_ver11
        tmp = re.sub('[<],[>]', '<,>', tmp)
        tmp = re.sub('[>],[<]', '>,<', tmp)
        tmp = re.sub('{[][<]', '{<', tmp)
        tmp = re.sub('[>][]}', '>}', tmp)
        tmp = re.sub('{[][$]', '{$', tmp)
        tmp = re.sub('[$][]}', '$}', tmp)
        return tmp
    
